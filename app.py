import streamlit as st
import pandas as pd
import gspread
from datetime import datetime
import re
import json
import google.generativeai as genai
from copy import copy
from openpyxl.utils import get_column_letter

# --- Page and PWA Configuration ---
st.set_page_config(
    page_title="Складов AI Асистент",
    layout="centered",
    page_icon="static/icon-192x192.png"  # Sets the browser tab icon
)
# Link to the PWA manifest
st.markdown('<link rel="manifest" href="/static/manifest.json">', unsafe_allow_html=True)


# --- API Key Configuration ---
# This function configures the Gemini API key from Streamlit Secrets
def configure_genai():
    try:
        api_key = st.secrets["GEMINI_API_KEY"]
        genai.configure(api_key=api_key)
        return True
    except (KeyError, FileNotFoundError):
        st.error("Вашият Gemini API ключ не е намерен в Streamlit Secrets. Моля, добавете го.")
        return False


# Configure the API at the start
AI_ENABLED = configure_genai()


# --- Google Sheets Authentication and Functions ---
@st.cache_resource
def get_gspread_client():
    """Initializes and returns the gspread client using Streamlit Secrets."""
    try:
        return gspread.service_account_from_dict(st.secrets["gcp_service_account"])
    except Exception as e:
        st.error(f"Грешка при свързване с Google Sheets: {e}")
        return None


@st.cache_data(ttl=60)  # Cache data for 60 seconds
def load_data_from_sheet(sheet_name):
    """Loads data from a specified Google Sheet into a pandas DataFrame."""
    client = get_gspread_client()
    if not client:
        return None
    try:
        sheet = client.open(sheet_name).sheet1
        header_row = 4 if sheet_name == "SalesData" else 1
        all_data = sheet.get_all_records(head=header_row, default_blank="")
        df = pd.DataFrame(all_data)
        df.columns = [re.sub(r'\s+', ' ', str(c).strip()) for c in df.columns]

        if 'Клиент име' in df.columns:
            df = df[df['Клиент име'].astype(str).str.strip().str.upper() != 'ОБЩО']

        return df.fillna('')
    except gspread.exceptions.SpreadsheetNotFound:
        st.error(f"Google Sheet с име '{sheet_name}' не е намерен. Моля, проверете името и правата за достъп.")
        return None
    except Exception as e:
        st.error(f"Грешка при зареждане на '{sheet_name}': {e}")
        return None


def update_inventory(product_name, quantity_to_subtract):
    """Finds a product in the Inventory sheet and updates its quantity."""
    client = get_gspread_client()
    if not client: return False
    try:
        sheet = client.open("Inventory").sheet1
        cell = sheet.find(product_name, in_column=1)
        if not cell:
            st.error(f"Продукт '{product_name}' не е намерен в 'Inventory'.")
            return False
        current_quantity = int(sheet.cell(cell.row, 2).value)
        new_quantity = current_quantity - quantity_to_subtract
        if new_quantity < 0:
            st.error(f"Недостатъчна наличност за '{product_name}'.")
            return False
        sheet.update_cell(cell.row, 2, new_quantity)
        load_data_from_sheet.clear()
        return True
    except Exception as e:
        st.error(f"Грешка при обновяване на инвентара: {e}")
        return False


def append_to_sales(row_data, all_cols):
    """Appends a new row and updates the SUM formulas in the 'Общо' row."""
    client = get_gspread_client()
    if not client: return False
    try:
        sheet = client.open("SalesData").sheet1

        # Find the "ОБЩО" row to insert above it
        total_cell = sheet.find("ОБЩО")
        insert_row_index = total_cell.row if total_cell else len(sheet.get_all_values()) + 1

        # Insert the new data row
        row_to_insert = [row_data.get(col, '') for col in all_cols]
        sheet.insert_row(
            row_to_insert,
            insert_row_index,
            value_input_option='USER_ENTERED',
            inherit_from_before=True
        )

        # --- NEW: LOGIC TO UPDATE SUM FORMULAS ---
        # 1. Determine the new location of the "Общо" row and the last data row
        new_total_row_index = insert_row_index + 1 if total_cell else -1
        last_data_row_index = new_total_row_index - 1

        if new_total_row_index != -1:
            # 2. Find the column range to update (from "Сума лв." to the column before "Цена")
            headers = sheet.row_values(4)  # Header is on row 4
            start_col_index = headers.index("Сума лв.") + 1
            end_col_index = headers.index("Цена")

            # 3. Create a list of formulas to update
            formulas_to_update = []
            # Calculate the full range of columns to sum, including newly added product columns
            num_product_cols = end_col_index - (headers.index("Общо кол-во") + 1)
            total_cols_to_sum = 1 + 1 + num_product_cols  # Suma lv + Obshto kol-vo + products

            # The range now goes from the start of sums to the end of the data columns
            effective_end_col_index = start_col_index + total_cols_to_sum

            for col_idx in range(start_col_index, effective_end_col_index):
                # We skip the columns that are not supposed to be summed (like product names in the total row)
                # Let's re-read the total row to see which cells have formulas
                total_row_values = sheet.row_values(new_total_row_index)
                if col_idx <= len(total_row_values) and str(total_row_values[col_idx - 1]).startswith('='):
                    col_letter = get_column_letter(col_idx)
                    # The formula sums from row 5 to the last data row
                    formula = f"=SUM({col_letter}5:{col_letter}{last_data_row_index})"
                    # Prepare the cell for batch update
                    cell_to_update = gspread.Cell(row=new_total_row_index, col=col_idx, value=formula)
                    formulas_to_update.append(cell_to_update)

            # 4. Update all formulas in one batch request for efficiency
            if formulas_to_update:
                sheet.update_cells(formulas_to_update, value_input_option='USER_ENTERED')

        load_data_from_sheet.clear()  # Clear cache
        return True
    except Exception as e:
        st.error(f"Грешка при добавяне на запис в 'SalesData': {e}")
        return False


# --- NEW: Function to automatically add a new product to Inventory ---
def add_new_product_to_inventory(product_name, initial_quantity):
    """Adds a new product and its quantity to the Inventory sheet."""
    client = get_gspread_client()
    if not client: return False, "Could not connect to Google Sheets."
    try:
        sheet = client.open("Inventory").sheet1
        if sheet.find(product_name, in_column=1):
            return False, f"Продукт с име '{product_name}' вече съществува в инвентара."
        sheet.append_row([product_name, initial_quantity], value_input_option='USER_ENTERED')
        load_data_from_sheet.clear()
        return True, f"Продуктът '{product_name}' е добавен успешно в 'Inventory'."
    except Exception as e:
        return False, f"Грешка при добавяне на нов продукт: {e}"


# --- NEW: Function to automatically add a new column to SalesData ---
def add_column_to_salesdata(product_name):
    """Safely inserts a new column for the product in the SalesData sheet."""
    client = get_gspread_client()
    if not client: return False, "Could not connect to Google Sheets."
    try:
        spreadsheet = client.open("SalesData")
        sheet = spreadsheet.sheet1

        # Find the column index of "Цена" (Price)
        headers = sheet.row_values(4)  # Header is on row 4
        try:
            price_col_index = headers.index("Цена") + 1
        except ValueError:
            return False, "Колона 'Цена' не е намерена в 'SalesData'."

        # Create the request body to insert a new column
        request = {
            "requests": [
                {
                    "insertDimension": {
                        "range": {
                            "sheetId": sheet.id,
                            "dimension": "COLUMNS",
                            "startIndex": price_col_index - 1,  # API is 0-indexed
                            "endIndex": price_col_index
                        },
                        "inheritFromBefore": True  # Inherit formatting from the column to the left
                    }
                }
            ]
        }

        # Execute the request
        spreadsheet.batch_update(body=request)

        # Update the header of the new column
        sheet.update_cell(4, price_col_index, product_name)
        load_data_from_sheet.clear()
        return True, f"Колоната '{product_name}' е добавена успешно в 'SalesData'."
    except Exception as e:
        return False, f"Грешка при добавяне на колона в 'SalesData': {e}"


# --- AI Search Function ---
def run_ai_doc_search(doc_number, data_string, doc_column_name):
    """Uses the improved 'few-shot' prompt to reliably find doc info."""
    if not AI_ENABLED:
        st.error("AI функцията е деактивирана поради липсващ API ключ.")
        return None
    model = genai.GenerativeModel('gemini-1.5-pro-latest')
    prompt = f"""
    You are an expert AI assistant for analyzing tabular data from a CSV string.
    Your task is to find all rows for document number '{doc_number}' and extract the specified information for each row into a valid JSON format.
    The main challenge is to correctly identify the 'Име на продукт' (Product Name) and 'Количество' (Quantity). The 'Име на продукт' is the *name of the column* that contains the quantity for that specific row. This column will be located between the 'Общо кол-во' and 'Цена' columns.
    ---
    **EXAMPLE:**
    *Input Data Snippet:*
    ```csv
    Клиент име,Бележка,Дата,Фактура №,Общо кол-во,Product A,Product B,Цена,Сума лв.
    ЗП ИВАН ПЕТРОВ,,2024-07-20,59460,10,,10,150.00,1500.00
    ```
    *Desired JSON Output for the example:*
    ```json
    {{
      "документи": [
        {{
          "Име на клиент": "ЗП ИВАН ПЕТРОВ",
          "Бележка": "",
          "Дата на издаване": "2024-07-20",
          "Име на продукт": "Product B",
          "Количество": "10",
          "Цена": "150.00",
          "Сума лв.": "1500.00"
        }}
      ]
    }}
    ```
    ---
    **NOW, ANALYZE THE FOLLOWING DATA AND PROVIDE THE JSON OUTPUT:**
    **Document to find:** '{doc_number}'
    **Data:**
    ```csv
    {data_string}
    ```
    **Instructions:**
    1. Find ALL rows where the '{doc_column_name}' column is exactly '{doc_number}'.
    2. For each matching row:
        - Extract "Име на клиент", "Бележка", "Дата", "Цена", and "Сума лв." directly from their columns. Use the 'Дата' value for "Дата на издаване".
        - To find "Име на продукт" and "Количество": Look at the columns between "Общо кол-во" and "Цена". The one column that has a number in it for this specific row is the "Име на продукт", and the number itself is the "Количество".
    3. If no document is found, return an empty JSON object like `{{"документи": []}}`.
    """
    try:
        response = model.generate_content(prompt)
        clean_response = response.text.strip().replace("```json", "").replace("```", "")
        if not clean_response or not clean_response.strip().startswith('{'):
            return None
        return json.loads(clean_response)
    except Exception:
        return None


# --- Main Streamlit App Logic ---
st.title("Складов AI Асистент 📦")

# Load data from Google Sheets
documents_df = load_data_from_sheet("SalesData")
inventory_df = load_data_from_sheet("Inventory")

app_mode = st.sidebar.radio(
    "Изберете режим на работа:",
    ("Добавяне на запис", "Добавяне на нов продукт", "Справка по Документ (с AI)", "Справка по Продукт")
)

# --- Mode 1: Add Entry ---
if app_mode == "Добавяне на запис":
    st.header("Добавяне на нов запис")
    if documents_df is None:
        st.warning("Не мога да заредя 'SalesData'. Проверете настройките.")
    else:
        all_cols = documents_df.columns.tolist()
        try:
            start_index = all_cols.index('Общо кол-во') + 1
            end_index = all_cols.index('Цена')
            product_list = all_cols[start_index:end_index]
        except (ValueError, IndexError):
            product_list = []
            st.error("Структурата на 'SalesData' е невалидна. Липсват 'Общо кол-во' или 'Цена'.")
        if product_list:
            with st.form("new_entry_form"):
                st.subheader("Данни за документа")
                client_name = st.text_input("Име на клиент")
                doc_number = st.text_input("Фактура №")
                doc_date = st.date_input("Дата на издаване", value=datetime.now())
                doc_note = st.text_input("Бележка")
                selected_product = st.selectbox("Избери продукт", options=product_list)
                quantity = st.number_input("Количество", min_value=1, step=1)
                price = st.number_input("Ед. цена (лв.)", min_value=0.0, format="%.2f")
                submitted = st.form_submit_button("✅ Запази записа")
                if submitted:
                    if not doc_number or not client_name:
                        st.warning("Моля, попълнете 'Фактура №' и 'Име на клиент'.")
                    else:
                        st.info("Обработка...")
                        if update_inventory(selected_product, quantity):
                            new_row_data = {col: '' for col in all_cols}
                            new_row_data['Дата'] = doc_date.strftime('%m/%d/%Y')
                            new_row_data['Фактура №'] = doc_number
                            new_row_data['Клиент име'] = client_name.upper()
                            new_row_data['Бележка'] = doc_note
                            new_row_data['Общо кол-во'] = int(quantity)
                            new_row_data[selected_product] = int(quantity)
                            new_row_data['Цена'] = float(price)
                            new_row_data['Сума лв.'] = float(quantity) * float(price)
                            if append_to_sales(new_row_data, all_cols):
                                st.success("✅ Записът е добавен и наличностите са обновени!")
                                st.balloons()
                            else:
                                st.error("❌ Грешка при запазване на записа. Проверете ръчно.")
                        else:
                            st.error("❌ Грешка при обновяване на инвентара.")

# --- UPDATED MODE: Add a New Product ---
elif app_mode == "Добавяне на нов продукт":
    st.header("Добавяне на нов продукт в системата")
    st.info("Тази форма автоматично ще добави продукта към 'Inventory' и ще добави нова колона в 'SalesData'.")

    with st.form("new_product_form"):
        new_product_name = st.text_input("Име на новия продукт")
        initial_quantity = st.number_input("Начална наличност", min_value=0, step=1)
        submitted = st.form_submit_button("✅ Добави продукт")

        if submitted:
            if not new_product_name:
                st.warning("Моля, въведете име на продукта.")
            else:
                with st.spinner("Добавяне на продукт..."):
                    # Step 1: Add to Inventory
                    inventory_success, inv_message = add_new_product_to_inventory(new_product_name, initial_quantity)

                    if inventory_success:
                        st.success(inv_message)
                        # Step 2: Add column to SalesData
                        sales_success, sales_message = add_column_to_salesdata(new_product_name)
                        if sales_success:
                            st.success(sales_message)
                            st.balloons()
                        else:
                            st.error(sales_message)
                            st.warning(
                                f"Продуктът '{new_product_name}' е добавен в Inventory, но колоната в SalesData не успя да се създаде. Моля, добавете я ръчно.")
                    else:
                        st.error(inv_message)

# --- Other Modes are unchanged ---
elif app_mode == "Справка по Документ (с AI)":
    st.header("Търсене по номер на документ (с AI)")
    if documents_df is None:
        st.error("Не мога да заредя данните от Google Sheets.")
    else:
        doc_column_name = 'Фактура №'
        if doc_column_name not in documents_df.columns:
            st.error(f"Липсва колона '{doc_column_name}' в 'SalesData'.")
        else:
            doc_number = st.text_input("Въведете номер на документ:").strip()
            if st.button("Търси с AI"):
                if not doc_number:
                    st.warning("Моля, въведете номер на документ.")
                else:
                    matching_docs_df = documents_df[documents_df[doc_column_name].astype(str) == doc_number]
                    if matching_docs_df.empty:
                        st.error(f"Документ №'{doc_number}' не е намерен.")
                    else:
                        data_string_subset = matching_docs_df.to_csv(index=False)
                        with st.spinner("AI анализира данните..."):
                            result = run_ai_doc_search(doc_number, data_string_subset, doc_column_name)
                        if not result or not result.get("документи"):
                            st.error(f"AI не успя да обработи документ №{doc_number}.")
                        else:
                            st.success(f"Намерени са {len(result['документи'])} записа за документ №{doc_number}")
                            for doc_item in result.get("документи", []):
                                with st.expander(
                                        f"**Клиент:** {doc_item.get('Име на клиент', '-')} | **Продукт:** {doc_item.get('Име на продукт', '-')}"):
                                    st.markdown(f"**Дата на издаване:** {doc_item.get('Дата на издаване', '-')}")
                                    st.markdown(f"**Бележка:** *{doc_item.get('Бележка', 'Няма')}*")
                                    st.markdown(
                                        f"**Количество:** {doc_item.get('Количество', '-')} | **Цена:** {doc_item.get('Цена', '-')} | **Сума:** {doc_item.get('Сума лв.', '-')}")

elif app_mode == "Справка по Продукт":
    st.header("Търсене по продукт")
    if documents_df is None or inventory_df is None:
        st.error("Не мога да заредя данните от Google Sheets.")
    else:
        all_cols = documents_df.columns.tolist()
        try:
            start_index = all_cols.index('Общо кол-во') + 1
            end_index = all_cols.index('Цена')
            product_list = all_cols[start_index:end_index]
            selected_product = st.selectbox("Изберете продукт:", product_list)
            if st.button("Търси"):
                if selected_product:
                    inventory_info = inventory_df[inventory_df['Продукт'] == selected_product]
                    quantity_available = inventory_info['Наличност'].iloc[0] if not inventory_info.empty else '0'
                    st.metric(label=f"Наличност за '{selected_product}'", value=f"{quantity_available} бр.")
                    matching_docs = documents_df[pd.to_numeric(documents_df[selected_product], errors='coerce').notna()]
                    st.subheader(f"Документи, съдържащи '{selected_product}':")
                    if matching_docs.empty:
                        st.info("Няма намерени документи за този продукт.")
                    else:
                        for index, row in matching_docs.iterrows():
                            with st.expander(
                                    f"**Документ №:** {row.get('Фактура №', '-')} | **Клиент:** {row.get('Клиент име', '-')}"):
                                st.markdown(f"**Дата:** {row.get('Дата', '-')}")
                                st.markdown(f"**Бележка:** *{row.get('Бележка', 'Няма')}*")
                                st.markdown(
                                    f"**Количество:** {row.get(selected_product, '-')} | **Цена:** {row.get('Цена', '-')} лв. | **Сума:** {row.get('Сума лв.', '-')} лв.")
        except (ValueError, IndexError):
            st.error("Структурата на 'SalesData' е невалидна.")